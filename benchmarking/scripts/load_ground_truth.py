#!/usr/bin/env python3
"""
Load ground-truth transcriptions from source_of_truth_transcriptions.xlsx.

Provides a helper used by compare_models.py and usable from a Colab notebook.
"""

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]

# SharePoint folder name → JSONL domain key (same mapping as sharepoint_batch_transcribe.py)
_DEPT_TO_DOMAIN: dict[str, str] = {
    "Grievance":  "grievance",
    "Helpdesk":   "helpdesk",
    "HR-Admin":   "hr_admin",
    "HR/Admin":   "hr_admin",
    "Impact":     "impact",
    "NextGen":    "nextgen",
    "Production": "production",
}


def load_ground_truth(
    xlsx_path: Path | None = None,
) -> dict[str, str]:
    """
    Read the xlsx and return {audio_key: human_reviewed_text}.

    audio_key has the form "helpdesk/HD_Q_2.ogg" to match the
    ``audio_file`` field in transcription JSONL files.

    Rows where "Human Reviewed" is empty are skipped.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("openpyxl is required: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    if xlsx_path is None:
        xlsx_path = ROOT / "source_of_truth_transcriptions.xlsx"

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    # Find column indices from header row
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    headers_lower = [str(h).strip().lower() if h else "" for h in headers]

    try:
        dept_idx = headers_lower.index("department")
    except ValueError:
        dept_idx = headers_lower.index("dept")
    fname_idx = headers_lower.index("file name")
    hr_idx = headers_lower.index("human reviewed")

    ground_truth: dict[str, str] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        department = str(row[dept_idx]).strip() if row[dept_idx] else ""
        filename = str(row[fname_idx]).strip() if row[fname_idx] else ""
        human_reviewed = str(row[hr_idx]).strip() if row[hr_idx] else ""

        if not filename or not human_reviewed:
            continue

        domain = _DEPT_TO_DOMAIN.get(department, department.lower().replace("-", "_"))
        key = f"{domain}/{filename}"
        ground_truth[key] = human_reviewed

    wb.close()
    return ground_truth


if __name__ == "__main__":
    gt = load_ground_truth()
    print(f"Loaded {len(gt)} ground-truth entries")
    for k, v in list(gt.items())[:3]:
        print(f"  {k}: {v[:60]}...")
