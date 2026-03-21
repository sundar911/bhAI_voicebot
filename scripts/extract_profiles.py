"""
Extract minimal bhAI user profiles from the Tiny Miracles employee workbook.

PRIVACY DESIGN
--------------
The source Excel contains ~65 columns of highly sensitive data (religion, caste,
disability, Aadhaar numbers, loan info, health history, etc.).

This script extracts ONLY the 10 columns needed for warm contextual conversation.
All other columns — including religion, caste, disability, loans, income, health —
are NEVER read or stored. The output profiles contain only safe context.

Profiles are encrypted with Fernet (BHAI_ENCRYPTION_KEY) before being written to
disk. They are gitignored and never included in version control.

Usage
-----
    # Set encryption key (from .env)
    export BHAI_ENCRYPTION_KEY=$(grep BHAI_ENCRYPTION_KEY .env | cut -d= -f2)

    # Run from project root
    python scripts/extract_profiles.py /path/to/workbook.xlsx [--dry-run]

Output: knowledge_base/users/+91XXXXXXXXXX.md (encrypted Fernet token)
"""

import argparse
import os
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: uv pip install openpyxl")
    sys.exit(1)

from src.bhai.security.crypto import encrypt_text


# ── Columns we extract (explicit allowlist — everything else is ignored) ─────

SAFE_COLUMNS = {
    "full_name":    "Full Name",
    "work_location":"Work Location",
    "work_type":    "Work Type",
    "employer_type":"Employer Type",
    "family_size":  "Family Size",
    "children":     "How many children are in your family? ",   # trailing space in source
    "children_age": "Children's Age",
    "children_pardhi": "How many children are in your family? Children's Age",  # Pardhi merged col
    "education":    "Highest Education Level",
    "eshram":       "Are you registered in E-Shram Scheme?",
    "skills":       "Type of Skill",
    # contact is extracted for filename only — never written to profile body
    "contact":      "Contact",
}

# Sheets to process (skip 'Data' which is just a combined view)
TARGET_SHEETS = ["MIDC", "Bombay Central", "Pardhi"]


# ── Helpers ───────────────────────────────────────────────────────────────────

def _clean(val) -> str:
    """Normalise a cell value to a clean string, or empty string if blank."""
    if val is None:
        return ""
    s = str(val).strip()
    # Non-breaking spaces and whitespace-only
    s = s.replace("\xa0", "").strip()
    if s.lower() in ("none", "n/a", "na", "-", "—"):
        return ""
    return s


def _first_name(full_name: str) -> str:
    """Extract only the first word of a full name."""
    parts = full_name.strip().split()
    return parts[0] if parts else full_name


def _normalise_contact(raw) -> str:
    """
    Normalise contact to +91XXXXXXXXXX format.
    Input may be an integer (e.g. 9324299326) or a string.
    Returns empty string if invalid.
    """
    try:
        digits = str(int(float(str(raw).replace(" ", "").replace("-", ""))))
    except (ValueError, TypeError):
        return ""

    # Remove +91 prefix if present
    if digits.startswith("91") and len(digits) == 12:
        digits = digits[2:]
    elif digits.startswith("0") and len(digits) == 11:
        digits = digits[1:]

    if len(digits) != 10:
        return ""

    return f"+91{digits}"


def _build_header_map(ws) -> dict:
    """Build {column_name: column_index} mapping from the first header row."""
    headers = [row for row in ws.iter_rows(min_row=1, max_row=1, values_only=True)][0]
    return {
        (str(h).strip() if h is not None else ""): idx
        for idx, h in enumerate(headers)
    }


def _get(row, header_map: dict, col_name: str) -> str:
    """Safely get a cell value from a row by column name."""
    idx = header_map.get(col_name)
    if idx is None:
        return ""
    val = row[idx] if idx < len(row) else None
    return _clean(val)


def _build_profile(row, header_map: dict, sheet_name: str) -> str:
    """
    Build a Hindi-language markdown profile from a single row.
    Only safe fields are included — all sensitive columns are ignored.
    """
    first_name      = _first_name(_get(row, header_map, SAFE_COLUMNS["full_name"]) or "")
    work_location   = _get(row, header_map, SAFE_COLUMNS["work_location"])
    work_type       = _get(row, header_map, SAFE_COLUMNS["work_type"])
    employer_type   = _get(row, header_map, SAFE_COLUMNS["employer_type"])
    family_size     = _get(row, header_map, SAFE_COLUMNS["family_size"])
    education       = _get(row, header_map, SAFE_COLUMNS["education"])
    eshram          = _get(row, header_map, SAFE_COLUMNS["eshram"])
    skills          = _get(row, header_map, SAFE_COLUMNS["skills"])

    # Pardhi sheet merges children count + age into one column
    if sheet_name == "Pardhi":
        children_info = _get(row, header_map, SAFE_COLUMNS["children_pardhi"])
        children_age  = ""
    else:
        children_info = _get(row, header_map, SAFE_COLUMNS["children"])
        children_age  = _get(row, header_map, SAFE_COLUMNS["children_age"])

    # ── Build profile sections ─────────────────────────────────────────────

    lines = ["# User Profile\n"]

    # Basics
    lines.append("## Basics")
    if first_name:
        lines.append(f"- Naam: {first_name}")
    if work_location:
        lines.append(f"- Workshop: {work_location}")

    work_desc_parts = [p for p in [work_type, employer_type] if p]
    if work_desc_parts:
        lines.append(f"- Kaam: {', '.join(work_desc_parts)}")
    if skills:
        lines.append(f"- Skills: {skills}")
    if education:
        lines.append(f"- Education: {education}")

    # Parivaar
    lines.append("\n## Parivaar")
    if family_size:
        lines.append(f"- Parivaar mein {family_size} log hain")

    if children_info:
        if children_age:
            lines.append(f"- Bachche: {children_info} — Ages: {children_age}")
        else:
            lines.append(f"- Bachche: {children_info}")
    elif not children_info:
        lines.append("- (Bachhon ki jaankari nahi hai)")

    # Schemes & Context
    lines.append("\n## Schemes & Work Context")
    eshram_status = "Registered" if eshram and eshram.lower() == "yes" else "Not registered"
    lines.append(f"- E-Shram: {eshram_status}")

    # Communication style + notes (left for human input)
    lines.append("\n## Communication Style")
    lines.append("- (Impact team: fill in after meeting them)")
    lines.append("\n## Notes for bhAI")
    lines.append("- (Impact team: add any relevant context here)")

    return "\n".join(lines)


# ── Main extraction logic ─────────────────────────────────────────────────────

def extract_profiles(excel_path: str, output_dir: Path, dry_run: bool = False) -> dict:
    """
    Extract and encrypt profiles from all target sheets.

    Returns a summary dict with counts per sheet.
    """
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    summary = {}
    total_written = 0
    total_skipped = 0

    output_dir.mkdir(parents=True, exist_ok=True)

    for sheet_name in TARGET_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: Sheet '{sheet_name}' not found, skipping")
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header_map = _build_header_map(ws)
        written = 0
        skipped = 0

        for row in rows[1:]:  # skip header
            # Skip entirely empty rows
            if all(v is None or str(v).strip() in ("", "\xa0") for v in row):
                continue

            raw_contact = row[header_map.get(SAFE_COLUMNS["contact"], -1)] if header_map.get(SAFE_COLUMNS["contact"]) is not None else None
            contact = _normalise_contact(raw_contact)

            if not contact:
                skipped += 1
                continue

            profile_text = _build_profile(row, header_map, sheet_name)
            output_path = output_dir / f"{contact}.md"

            if not dry_run:
                encrypted = encrypt_text(profile_text)
                output_path.write_text(encrypted, encoding="utf-8")
            else:
                print(f"\n[DRY RUN] Would write: {output_path.name}")
                print("  Preview (first 5 lines of plaintext):")
                for line in profile_text.splitlines()[:5]:
                    print(f"    {line}")
            written += 1

        summary[sheet_name] = {"written": written, "skipped": skipped}
        total_written += written
        total_skipped += skipped
        print(f"  {sheet_name}: {written} profiles {'previewed' if dry_run else 'written'}, {skipped} skipped (no contact)")

    print(f"\nTotal: {total_written} profiles {'previewed' if dry_run else 'written'}, {total_skipped} skipped")
    return summary


def main():
    parser = argparse.ArgumentParser(
        description="Extract minimal bhAI profiles from Tiny Miracles employee workbook."
    )
    parser.add_argument("excel_path", help="Path to the Excel workbook (keep OUTSIDE project dir)")
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Preview profiles without writing or encrypting anything"
    )
    args = parser.parse_args()

    excel_path = Path(args.excel_path)
    if not excel_path.exists():
        print(f"ERROR: File not found: {excel_path}")
        sys.exit(1)

    # Safety check: make sure the Excel isn't inside the project folder
    try:
        excel_path.resolve().relative_to(ROOT)
        print("WARNING: The Excel file appears to be inside the project directory.")
        print("  Best practice: keep it in ~/Downloads/ or another location outside the project.")
        response = input("  Continue anyway? [y/N]: ").strip().lower()
        if response != "y":
            sys.exit(0)
    except ValueError:
        pass  # Good — it's outside the project dir

    if not args.dry_run and not os.environ.get("BHAI_ENCRYPTION_KEY"):
        print("ERROR: BHAI_ENCRYPTION_KEY environment variable is not set.")
        print("  Set it with: export BHAI_ENCRYPTION_KEY=$(grep BHAI_ENCRYPTION_KEY .env | cut -d= -f2)")
        sys.exit(1)

    output_dir = ROOT / "knowledge_base" / "users"
    mode = "DRY RUN — no files will be written" if args.dry_run else f"Output → {output_dir}"

    print(f"\nbhAI Profile Extractor")
    print(f"  Source: {excel_path}")
    print(f"  Mode: {mode}")
    print(f"  Extracting 10 safe columns (all sensitive data excluded)\n")

    extract_profiles(str(excel_path), output_dir, dry_run=args.dry_run)

    if not args.dry_run:
        print(f"\nProfiles are encrypted with Fernet.")
        print(f"They are gitignored — do NOT commit them.")
        print(f"\nSecurity reminder: delete the Excel file from your local machine once done:")
        print(f"  rm '{excel_path}'")


if __name__ == "__main__":
    main()
